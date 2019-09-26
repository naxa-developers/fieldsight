from __future__ import absolute_import
import time
import re 
import os
import json
import datetime
import gc
from datetime import date
from django.db import transaction
from django.contrib.gis.geos import Point
from celery import shared_task
from onadata.apps.fieldsight.models import Organization, Project, Site, Region, SiteType, ProjectType, ProgressSettings, SiteMetaAttrAnsHistory
from onadata.apps.fieldsight.utils.progress import set_site_progress
from onadata.apps.userrole.models import UserRole
from onadata.apps.eventlog.models import FieldSightLog, CeleryTaskProgress
from channels import Group as ChannelGroup
from django.contrib.auth.models import User, Group
from onadata.apps.fieldsight.fs_exports.formParserForExcelReport import parse_form_response
from io import BytesIO
from django.shortcuts import get_object_or_404
from onadata.apps.fsforms.models import FieldSightXF, FInstance, Stage
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db.models import Prefetch
from .generatereport import PDFReport
import os, tempfile, zipfile, dateutil
from django.conf import settings

from django.http import HttpResponse
from django.core.servers.basehttp import FileWrapper
from openpyxl import Workbook

from openpyxl.styles import Font

from django.core.files.storage import get_storage_class
from onadata.libs.utils.viewer_tools import get_path
from PIL import Image
import tempfile, zipfile

from onadata.libs.utils.viewer_tools import get_path
import pyexcel as p
from .metaAttribsGenerator import get_form_answer, get_form_sub_status, get_form_submission_count, get_form_ques_ans_status
from django.conf import settings
from django.db.models import Sum, Case, When, IntegerField, Count
from django.core.exceptions import MultipleObjectsReturned
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.utils import timezone

from dateutil.rrule import rrule, MONTHLY, DAILY
from django.db import connection                                         
from onadata.apps.logger.models import Instance, Attachment
from onadata.apps.fieldsight.fs_exports.log_generator import log_types

from django.db.models import Q
from django.contrib.contenttypes.models import ContentType

from onadata.apps.fsforms.reports_util import get_images_for_site_all
from onadata.apps.users.signup_tokens import account_activation_token
from onadata.apps.subscriptions.models import Subscription, Package, TrackPeriodicWarningEmail
from onadata.apps.fsforms.models import InstanceStatusChanged

from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

form_status_map=["Pending", "Rejected", "Flagged", "Approved"]

def cleanhtml(raw_html):
   cleanr = re.compile('<\S.*?>')
   cleantext = re.sub(cleanr, '', raw_html)
   return cleantext

class DriveException(Exception):
    pass

@shared_task()
def gsuit_assign_perm(title, emails):
    time.sleep(5)
    try:
        gauth = GoogleAuth()
        drive = GoogleDrive(gauth)
        file = drive.ListFile({'q':"title = '"+ title +"' and trashed=false"}).GetList()[0]
        for perm in emails:
            time.sleep(1)
            file.InsertPermission({
                'type':'user',
                'value':perm,
                'role': 'writer'
            })
    except:
        pass


def upload_to_drive(file_path, title, folder_title, project, user):
    # pass
    """ TODO: folder names of 'Site Details' and 'Site Progress' must be in google drive."""
    try:
        gauth = GoogleAuth()
        drive = GoogleDrive(gauth)

        folders = drive.ListFile({'q':"title = '"+ folder_title +"'"}).GetList()
        
        if folders:
            folder_id = folders[0]['id']
        else:
            folder_metadata = {'title' : folder_title, 'mimeType' : 'application/vnd.google-apps.folder'}
            new_folder = drive.CreateFile(folder_metadata)
            new_folder.Upload()            
            folder_id = new_folder['id']
        
        file = drive.ListFile({'q':"title = '"+ title +"' and trashed=false"}).GetList()

        if not file:    
            new_file = drive.CreateFile({'title' : title, "parents": [{"kind": "drive#fileLink", "id": folder_id}]})
            new_file.SetContentFile(file_path)
            new_file.Upload({'convert':True})
            file = drive.ListFile({'q':"title = '"+ title +"' and trashed=false"}).GetList()[0]

        else:
            file = file[0]
            file.SetContentFile(file_path)
            file.Upload({'convert':True})
        
        _project = Project.objects.get(pk=project.id) 
        gsuit_meta = _project.gsuit_meta
        gsuit_meta[folder_title] = {
            'link':file['alternateLink'],
            'updated_at':datetime.datetime.now().isoformat(),
        }
        if user:
            gsuit_meta[folder_title]['user'] = {
                'username': user.username,
                'full_name': user.get_full_name()
            }
        _project.gsuit_meta = gsuit_meta
        _project.save()
        permissions = file.GetPermissions()

        user_emails = _project.project_roles.filter(group__name__in=["Project Manager", "Project Donor"], ended_at__isnull = True, site=None).distinct('user').values_list('user__email', flat=True)
        
        all_users = set(user_emails)

        existing_perms = []

        for permission in permissions:
            existing_perms.append(permission['emailAddress'])

        perms = set(existing_perms)

        perm_to_rm = perms - all_users
        perm_to_add = all_users - perms

        for permission in permissions:
            if permission['emailAddress'] in perm_to_rm and permission['emailAddress'] != "exports.fieldsight@gmail.com":
                file.DeletePermission(permission['id'])

        try:
            index = 0
            for perm in perm_to_add:
                file.InsertPermission({
                            'type':'user',
                            'value':perm,
                            'role': 'writer'
                        })
                index += 1
        except:
            gsuit_assign_perm.delay(title, perm[index:])

    except Exception as e:
        raise DriveException({"message":e})



@shared_task()
def site_download_zipfile(task_prog_obj_id, size):
    time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status = 1
    task.save()
    try:
        default_storage = get_storage_class()() 
        buffer = BytesIO()
        datas = get_images_for_site_all(task.object_id)
        urls = list(datas["result"])
        archive = zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED)
        index=0
        username=urls[0]['_attachments']['download_url'].split('/')[2]
        for url in urls:        
            index+=1
            if default_storage.exists(get_path(url['_attachments']['filename'], size)):
                
                with tempfile.NamedTemporaryFile(mode="wb") as temp:
                
                    file = default_storage.open(get_path(url['_attachments']['filename'], size))
                    img=Image.open(file)
                    img.save(temp, img.format)
                    # filename = '/srv/fieldsight/fieldsight-kobocat'+url['_attachments']['filename'] # Select your files here.                           
                    archive.write(temp.name, url['_attachments']['filename'].split('/')[2])
                    
        archive.close()
        buffer.seek(0)
        zipFile = buffer.getvalue()
        if default_storage.exists(task.content_object.identifier + '/files/'+task.content_object.name+'.zip'):
            default_storage.delete(task.content_object.identifier + '/files/'+task.content_object.name+'.zip')
        zipFile_url = default_storage.save(task.content_object.identifier + '/files/'+task.content_object.name+'.zip', ContentFile(zipFile))
        task.file.name = zipFile_url
        task.status = 2
        task.save()
        buffer.close()
        noti = task.logs.create(source=task.user, type=32, title="Image Zip generation in site",
                                   recipient=task.user, content_object=task, extra_object=task.content_object,
                                   extra_message=" <a href='"+ task.file.url +"'>Image Zip file </a> generation in site")
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Report Gen Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Image Zip generation in site",
                                       content_object=task.content_object, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))
        buffer.close()                                                                      

@shared_task(time_limit=300, soft_time_limit=300)
def generate_stage_status_report(task_prog_obj_id, project_id, site_type_ids, region_ids, sync_to_drive=False):
    time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    project = Project.objects.get(pk=project_id)
    task.status = 1
    task.save()
    try:
        data=[]
        ss_index = []
        form_ids = []
        stages_rows = []
        head_row = ["Site ID", "Name", "Region ID", "Address", "Latitude", "longitude", "Status"]

        query={}
        
        stages = project.stages.filter(stage__isnull=True)
        for stage in stages:
            sub_stages = stage.parent.all()
            if len(sub_stages):
                head_row.append("Stage :"+stage.name)
                stages_rows.append("Stage :"+stage.name)
                ss_index.append(str(""))
                for ss in sub_stages:
                    head_row.append("Sub Stage :"+ss.name)
                    ss_index.append(str(ss.stage_forms.id))
                    form_ids.append(str(ss.stage_forms.id))
                    query[str(ss.stage_forms.id)] = Sum(
                        Case(
                        When(site_instances__project_fxf_id=ss.stage_forms.id, then=1),
                        default=0, output_field=IntegerField()
                        ))

        query['flagged'] = Sum(
            Case(
                When(site_instances__form_status=2, site_instances__project_fxf_id__in=form_ids, then=1),
                default=0, output_field=IntegerField()
            ))

        query['rejected'] = Sum(
            Case(
                When(site_instances__form_status=1, site_instances__project_fxf_id__in=form_ids, then=1),
                default=0, output_field=IntegerField()
            ))
         
        query['submission'] = Sum(
            Case(
                When(site_instances__project_fxf_id__in=form_ids, then=1),
                default=0, output_field=IntegerField()
            ))

        head_row.extend(["Site Visits", "Submission Count", "Flagged Submission", "Rejected Submission"])
        data.append(head_row)
        
        sites = Site.objects.filter(is_active=True)

        sites_filter = {'project_id': project.id}
        finstance_filter = {'project_fxf__in': form_ids}
        
        if site_type_ids:
            sites_filter['type_id__in'] = site_type_ids
            finstance_filter['site__type_id__in'] = site_type_ids

        if region_ids:
            sites_filter['region_id__in']=region_ids
            finstance_filter['site_id__in'] = site_type_ids

        site_dict = {}


        # Redoing query because annotate and lat long did not go well in single query.
        # Probable only an issue because of old django version.

        
        for site_obj in sites.filter(**sites_filter).iterator():
            site_dict[str(site_obj.id)] = {'visits':0,'site_status':'No Submission', 'latitude':site_obj.latitude,'longitude':site_obj.longitude}

        sites_status=FInstance.objects.filter(**finstance_filter).order_by('site_id','-id').distinct('site_id').values_list('site_id', 'form_status')
        
        for site_status in sites_status:
            try:
                site_dict[str(site_status[0])]['site_status'] = form_status_map[site_status[1]]
            except:
                pass
        sites_status = None
        gc.collect()
        
        site_visits = settings.MONGO_DB.instances.aggregate([{"$match":{"fs_project": project.id, "fs_project_uuid": {"$in":form_ids}}},  { "$group" : { 
              "_id" :  { 
                "fs_site": "$fs_site",
                "date": { "$substr": [ "$start", 0, 10 ] }
              },
           }
         }, { "$group": { "_id": "$_id.fs_site", "visits": { 
                  "$push": { 
                      "date":"$_id.date"
                  }          
             }
         }}])['result']

        for site_visit in site_visits:
            try:
                site_dict[str(site_visit['_id'])]['visits'] = len(site_visit['visits'])
            except:
                pass

        site_visits = None
        gc.collect()

        
        sites = sites.filter(**sites_filter).values('id','identifier', 'name', 'region__identifier', 'address').annotate(**query)
        
        for site in sites:
            # import pdb; pdb.set_trace();
            try:
                site_row = [site['identifier'], site['name'], site['region__identifier'], site['address'], site_dict[str(site.get('id'))]['latitude'], site_dict[str(site.get('id'))]['longitude'], site_dict[str(site.get('id'))]['site_status']]
                
                for stage in ss_index:
                    site_row.append(site.get(stage, ""))

                site_row.extend([site_dict[str(site.get('id'))]['visits'], site['submission'], site['flagged'], site['rejected']])

                data.append(site_row)
            except Exception as e:
                print e

        sites = None
        site_dict = None
        gc.collect()

        p.save_as(array=data, dest_file_name="media/stage-report/{}_stage_data.xls".format(project.id))
        
        with open("media/stage-report/{}_stage_data.xls".format(project.id), 'rb') as fin:
            buffer = BytesIO(fin.read())
            buffer.seek(0)
            path = default_storage.save(
                "media/stage-report/{}_stage_data.xls".format(project.id),
                ContentFile(buffer.getvalue())
            )
            buffer.close()

        task.file.name = path
        task.status = 2
        task.save()
        
        if sync_to_drive:
            upload_to_drive("media/stage-report/{}_stage_data.xls".format(project.id), "{} - Progress Report".format(project.id), "Site Progress", project, task.user)

            noti = task.logs.create(source=task.user, type=32, title="Site Stage Progress report sync to Google Drive",
                                   recipient=task.user, content_object=project, extra_object=project,
                                   extra_message="Site Stage Progress report sync to Google Drive in project")


        else:
            noti = task.logs.create(source=task.user, type=32, title="Site Stage Progress report generation in Project",
                                   recipient=task.user, content_object=project, extra_object=project,
                                   extra_message=" <a href='/"+ "media/stage-report/{}_stage_data.xls".format(project.id) +"'>Site Stage Progress report </a> generation in project")

    except DriveException as e:
        print 'Report upload to drive  Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Site Stage Progress report upload to Google Drive in Project",
                                       content_object=project, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))

    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Report Gen Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Site Stage Progress report generation in Project",
                                       content_object=project, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))

     
@shared_task()
def UnassignUser(task_prog_obj_id, user_id, sites, regions, projects, group_id):
    time.sleep(5)
    user = User.objects.get(pk=user_id)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status=1
    task.save()
    
    try:
        count = 0
        with transaction.atomic():
            
            if sites:
                
                for site_id in sites:
                    roles=UserRole.objects.filter(user_id=user_id, site_id = site_id, group_id = group_id, ended_at=None)
                    for role in roles:
                        role.ended_at = datetime.datetime.now()
                        role.save()
                        count = count + 1


            if regions:
                for region_id in regions:
                    sites = Site.objects.filter(region_id=region_id[1:])    
                    
                    for site_id in sites:
                        roles=UserRole.objects.filter(user_id=user_id, site_id = site_id, group_id = group_id, ended_at=None)
                        for role in roles:
                            role.ended_at = datetime.datetime.now()
                            role.save()
                            count = count + 1

            if projects:
                for project_id in projects: 
                    sites = Site.objects.filter(project_id = project_id[1:])    
                    for site_id in sites:
                        roles=UserRole.objects.filter(user_id=user_id, site_id = site_id, group_id = group_id, ended_at=None)
                        for role in roles:
                            role.ended_at = datetime.datetime.now()
                            role.save()
                            count = count + 1

            task.status = 2
            task.save()
            if group_id == "3":
                extra_message= "removed " + str(count) + "Reviewer Roles"
            else:
                extra_message= "removed " + str(count) + " Supervisor Roles"

            noti = task.logs.create(source=task.user, type=35, title="Remove Roles",
                                       content_object=user.user_profile, recipient=task.user,
                                       extra_message=extra_message)
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Role Remove Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Role Remove for ",
                                       content_object=user.user_profile, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))


@shared_task()
def UnassignAllProjectRolesAndSites(task_prog_obj_id, project_id):
    time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status=1
    task.save()
    project = Project.all_objects.get(pk=project_id)
    try:
        
        sites_count = 0
        roles_count = 0

        with transaction.atomic():        
            roles=UserRole.objects.filter(project_id = project_id, ended_at=None)
            for role in roles:
                role.ended_at = datetime.datetime.now()
                role.save()
                roles_count = roles_count + 1
   
            sites=Site.objects.filter(project_id = project_id)
            for site in sites:
                site.is_active = False
                site.save()
                sites_count = sites_count + 1

            task.status = 2
            task.save()
            
            extra_message= "removed " + str(roles_count) + " User Roles and " + str(sites_count) + " sites "

            
            noti = task.logs.create(source=task.user, type=35, title="Remove Roles",
                                           content_object=project, recipient=task.user,
                                           extra_message=extra_message)
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Role Remove Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Role Remove for ",
                                       content_object=project, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))


@shared_task()
def UnassignAllSiteRoles(task_prog_obj_id, site_id):
    time.sleep(5)
    site = Site.all_objects.get(pk=site_id)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status=1
    task.save()
    
    try:
        count = 0
        with transaction.atomic():        
            roles=UserRole.objects.filter(site_id = site_id, ended_at=None)
            for role in roles:
                role.ended_at = datetime.datetime.now()
                role.save()
                count = count + 1

            task.status = 2
            task.save()
            
            extra_message= "removed " + str(count) + " User Roles "

            noti = task.logs.create(source=task.user, type=35, title="Remove Roles",
                                       content_object=site, recipient=task.user,
                                       extra_message=extra_message)
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Role Remove Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Role Remove for ",
                                       content_object=site, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))


def get_site_type(value):
    try:
        return int(value)
    except:
        return 0    

@shared_task()
def bulkuploadsites(task_prog_obj_id, sites, pk):
    time.sleep(2)
    project = Project.objects.get(pk=pk)
    # task_id = bulkuploadsites.request.id
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.content_object = project
    task.status=1
    task.save()
    count = ""
    try:
        sites
        count = len(sites)
        task.description = "Bulk Upload of "+str(count)+" Sites."
        task.save()
        new_sites = 0
        updated_sites = 0
        with transaction.atomic():
            i=0
            interval = count/20
            for site in sites:
                # time.sleep(0.7)
                print(i)
                site = dict((k, v) for k, v in site.iteritems() if v is not '')

                lat = site.get("longitude", 85.3240)
                long = site.get("latitude", 27.7172)
                
                if lat == "":
                    lat = 85.3240
                if long == "":
                    long = 27.7172

                location = Point(round(float(lat), 6), round(float(long), 6), srid=4326)
                region_idf = site.get("region_id", None)
                

                type_identifier = site.get("type", None)

                _site, created = Site.objects.get_or_create(identifier=str(site.get("identifier")),
                                                                project=project)

                if created:
                    new_sites += 1
                else:
                    updated_sites += 1

                if type_identifier:
                     site_type = SiteType.objects.get(identifier=type_identifier, project=project)
                     _site.type = site_type
                
                region = None
                
                if region_idf is not None:
                    region = Region.objects.get(identifier=str(region_idf), project = project)
                        
                _site.region = region
                _site.name = site.get("name")
                _site.phone = site.get("phone")
                _site.address = site.get("address")
                _site.public_desc = site.get("public_desc")
                _site.additional_desc = site.get("additional_desc")
                _site.location = location
                # _site.logo = "logo/default_site_image.png"

                meta_ques = project.site_meta_attributes

                myanswers = _site.site_meta_attributes_ans
                for question in meta_ques:
                    if question['question_type'] not in ['Form','FormSubStat','FormSubCountQuestion','FormQuestionAnswerStatus']:
                        myanswers[question['question_name']]=site.get(question['question_name'], "")
                
                _site.site_meta_attributes_ans = myanswers
                _site.save()
                i += 1
                
                if i > interval:
                    interval = i+interval
                    bulkuploadsites.update_state('PROGRESS', meta={'current': i, 'total': count})
            task.status = 2
            task.save()

            extra_message= ""
            if new_sites > 0 and updated_sites > 0:
                extra_message = " updated " + str(updated_sites) + " Sites and" + " created " + str(new_sites) + " Sites"
            elif new_sites > 0 and updated_sites == 0:
                extra_message = " created " + str(new_sites) + " Sites"
            elif new_sites == 0 and updated_sites > 0:
                extra_message = " updated " + str(updated_sites) + " Sites"
            

            noti = project.logs.create(source=task.user, type=12, title="Bulk Sites",
                                       organization=project.organization,
                                       project=project, content_object=project, extra_object=project,
                                       extra_message=extra_message)
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Site Upload Unsuccesfull. %s' % e
        print e.__dict__
        noti = project.logs.create(source=task.user, type=412, title="Bulk Sites",
                                       content_object=project, recipient=task.user,
                                       extra_message=str(count) + " Sites @error " + u'{}'.format(e.message))
        

@shared_task()
def generateCustomReportPdf(task_prog_obj_id, site_id, base_url, fs_ids, start_date, end_date, removeNullField):
    time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status = 1
    site=get_object_or_404(Site, pk=site_id)
    task.content_object = site
    task.save()

    try:
        buffer = BytesIO()
        report = PDFReport(buffer, 'Letter')
        pdf = report.generateCustomSiteReport(site_id, base_url, fs_ids, start_date, end_date, removeNullField)
        
        buffer.seek(0)
        pdf = buffer.getvalue()
        pdf_url = default_storage.save(site.name + '/pdf/'+site.name+'-submissions.pdf', ContentFile(pdf))
        buffer.close()
        task.file.name = pdf_url

        task.status = 2
        task.save()

        noti = task.logs.create(source=task.user, type=32, title="Pdf Report generation in site",
                                   recipient=task.user, content_object=task, extra_object=site,
                                   extra_message=" <a href='"+ task.file.url +"'>Pdf report</a> generation in site")
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Report Gen Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Pdf Report generation in site",
                                       content_object=site, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))
        buffer.close()        

def siteDetailsGenerator(project, sites, ws):
    try:
        header_columns = [ {'id': 'identifier' ,'name':'identifier'},
                           {'id': 'name','name':'name'},
                           {'id': 'site_type_identifier','name':'type'}, 
                           {'id': 'phone','name':'phone'},
                           {'id': 'address','name':'address'},
                           {'id': 'public_desc','name':'public_desc'},
                           {'id': 'additional_desc','name':'additional_desc'},
                           {'id': 'latitude','name':'latitude'},
                           {'id': 'longitude','name':'longitude'}, ]
        
        if project.cluster_sites:
            header_columns += [{'id':'region_identifier', 'name':'region_id'}, ]
        
        meta_ques = project.site_meta_attributes
        for question in meta_ques:
            header_columns += [{'id': question['question_name'], 'name':question['question_name']}]
        
        
        get_answer_questions = []
        get_sub_count_questions = []
        get_sub_status_questions = []   
        get_answer_status_questions = []

        site_list = {}
        meta_ref_sites = {}
        site_submission_count = {}
        site_sub_status = {}

        
        for meta in meta_ques: 
            if meta['question_type'] == 'FormSubStat':
                get_sub_status_questions.append(meta)

            elif meta['question_type'] == 'FormSubCountQuestion':
                get_sub_count_questions.append(meta)


        if get_sub_count_questions:
            query = {}
            for meta in get_sub_count_questions:
                query[meta['question_name']] = Sum(
                                        Case(
                                            When(site_instances__project_fxf_id=meta['form_id'], then=1),
                                            default=0, output_field=IntegerField()
                                        ))
            results = sites.values('id',).annotate(**query)
            for submission_count in results:
                site_submission_count[submission_count['id']] = submission_count

        if get_sub_status_questions:
            query = {}
            for meta in get_sub_status_questions:
                for submission in FInstance.objects.filter(project_id=project.id, project_fxf_id=meta['form_id']).values('site_id', 'date').distinct('site_id').order_by('site_id', '-instance_id'):
                    try:
                        site_sub_status[meta['form_id']][submission['site_id']] = "Last submitted on " + submission['date'].strftime("%d %b %Y %I:%M %P")
                    except:
                        site_sub_status[meta['form_id']] = {submission['site_id']:"Last submitted on " + submission['date'].strftime("%d %b %Y %I:%M %P")}

        #Optimized query, only one query per link type meta attribute which covers all site's answers.

        def generate(project_id, site_map, meta, identifiers, selected_metas):
            project_id = str(project_id)
            sub_meta_ref_sites = {}
            sub_site_map = {}
            
            sitesnew = Site.objects.filter(identifier__in = identifiers, project_id = project_id)
            
            for site in sitesnew.iterator():
                if project_id == str(project.id):
                    continue
            
                identifier = site_map.get(site.identifier)
                  
                if not site.site_meta_attributes_ans:
                    meta_ans = {}
                else:
                    meta_ans = site.site_meta_attributes_ans

                for meta in selected_metas.get(project_id, []):
                    
                    if meta.get('question_type') == "Link":
                        link_answer=str(meta_ans.get(meta.get('question_name'), ""))
                        if link_answer != "":    
                            if meta['question_name'] in sub_site_map:
                                if site.identifier in sub_site_map[meta['question_name']]:
                                    sub_site_map[meta['question_name']][link_answer].append(identifier)
                                else:
                                    sub_site_map[meta['question_name']][link_answer] = [identifier]
                            else:
                                sub_site_map[meta['question_name']] = {}
                                sub_site_map[meta['question_name']][link_answer] = [identifier]
                            
                            for idf in identifier:
                                if meta['question_name'] in sub_meta_ref_sites:
                                    sub_meta_ref_sites[meta['question_name']].append(meta_ans.get(meta['question_name']))
                                else:
                                    sub_meta_ref_sites[meta['question_name']] = [meta_ans.get(meta['question_name'])]

                    else:
                        for idf in identifier:
                            site_list[idf][project_id+"-"+meta.get('question_name')] = meta_ans.get(meta.get('question_name'), "")
            
            del sitesnew
            gc.collect()

            for meta in selected_metas.get(project_id, []):
                head = header_columns
                head += [{'id':project_id+"-"+meta.get('question_name'), 'name':meta.get('question_text')}]
                if meta.get('question_type') == "Link":
                    generate(meta['project_id'], sub_site_map.get(meta['question_name'], []), meta, sub_meta_ref_sites.get(meta['question_name'], []), selected_metas)


        for site in sites.select_related('region').iterator():
            
            columns = {'identifier':site.identifier, 'name':site.name, 'site_type_identifier':site.type.identifier if site.type else "", 'phone':site.phone, 'address':site.address, 'public_desc':site.public_desc, 'additional_desc':site.additional_desc, 'latitude':site.latitude,
                       'longitude':site.longitude, }
            
            if project.cluster_sites:
                columns['region_identifier'] = site.region.identifier if site.region else ""
            
            meta_ques = project.site_meta_attributes
            meta_ans = site.site_meta_attributes_ans
            for question in meta_ques:
                if question['question_type'] == 'FormSubCountQuestion':
                    columns[question['question_name']] = site_submission_count[site.id][question['question_name']]
                elif question['question_type'] == 'FormSubStat':
                    columns[question['question_name']] = site_sub_status[question['form_id']].get(site.id, '') if question['form_id'] in site_sub_status else ''
                elif question['question_type'] in ['Form','FormQuestionAnswerStatus']:
                    columns[question['question_name']] = ""

                else:
                    if question['question_name'] in meta_ans:
                        columns[question['question_name']] = meta_ans[question['question_name']]

                        if question['question_type'] == "Link" and meta_ans[question['question_name']] != "":
                            if question.get('question_name') in meta_ref_sites:
                                meta_ref_sites[question.get('question_name')].append(meta_ans[question['question_name']])
                            else:
                                meta_ref_sites[question.get('question_name')] = [meta_ans[question['question_name']]]
                    
                    else:
                        columns[question['question_name']] = '-'
            
            site_list[site.id] = columns
        
        del sites
        gc.collect()

        for meta in meta_ques:
            if meta['question_type'] == "Link":
                site_map = {}
                for key, value in site_list.items():
                    if value[meta['question_name']] != "":
                        identifier = str(value.get(meta['question_name']))
                        if identifier in site_map:
                            site_map[identifier].append(key)
                        else:
                            site_map[identifier] = [key]
                
                generate(meta['project_id'], site_map, meta, meta_ref_sites.get(meta['question_name'], []), meta.get('metas'))

            elif meta['question_type'] == 'Form':
                get_answer_questions.append(meta)

            elif meta['question_type'] == 'FormQuestionAnswerStatus':
                get_answer_status_questions.append(meta)
                    
        for meta in get_answer_questions:
            form_owner = None
            query = settings.MONGO_DB.instances.aggregate([
                {"$sort":{"_id":1}},
                {"$match":{"fs_project": project.id, "fs_project_uuid": str(meta['form_id'])}},  { "$group" : {
                "_id" : "$fs_site",
                "answer": { '$last': "$"+meta['question']['name'] }
               }
             }])

            for submission in query['result']:
                try:    
                    if meta['question']['type'] in ['photo', 'video', 'audio'] and submission['answer'] is not "":
                        if not form_owner:
                            form_owner = FieldSightXF.objects.select_related('xf__user').get(pk=meta['form_id']).xf.user.username
                        site_list[int(submission['_id'])][meta['question_name']] = 'http://app.fieldsight.org/attachment/medium?media_file='+  +'/attachments/'+submission['answer']
                    
                    if meta['question']['type'] == 'repeat':
                        site_list[int(submission['_id'])][meta['question_name']] = ""

                    site_list[int(submission['_id'])][meta['question_name']] = submission['answer']
                except:
                    pass

        for meta in get_answer_status_questions:
        
            query = settings.MONGO_DB.instances.aggregate([
                {"$sort": {"_id": 1}},
                {"$match":{"fs_project": project.id, "fs_project_uuid": str(meta['form_id'])}},  { "$group" : {
                "_id" : "$fs_site",
                "answer": { '$last': "$"+meta['question']['name'] }
               }
             }])

            for submission in query['result']:
                try:
                    if submission['answer'] and submission['answer'] != "":
                        site_list[int(submission['_id'])][meta['question_name']] = "Answered"
                    else:
                        site_list[int(submission['_id'])][meta['question_name']] = "Not Answered"
                except:
                    pass
                    
        row_num = 0
        
        header_row=[]
        for col_num in range(len(header_columns)):
            # header_cell=WriteOnlyCell(ws, value=header_columns[col_num]['name'])
            # header_cell=Font(name='Courier', size=16)
            header_row.append(header_columns[col_num]['name'])
            
        ws.append(header_row)
        


        for key,site in site_list.iteritems():
        #    ws.append([site.get(header_columns[col_num]['id']) for col_num in range(len(header_columns))])
            row=[]
            for col_num in range(len(header_columns)):
                row.append(site.get(header_columns[col_num]['id'], ""))    
            ws.append(row)

        gc.collect()
        return True, 'success'

    except Exception as e:
        gc.collect()
        return False, e.message

# project = Project.objects.get(pk=137)
# sites = project.sites.all()
# siteDetailsGenerator(project, sites, None)

@shared_task(time_limit=300, soft_time_limit=300)
def generateSiteDetailsXls(task_prog_obj_id, project_id, region_ids, type_ids=None, sync_to_drive=False):
    time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status = 1
    project = get_object_or_404(Project, pk=project_id)
    task.content_object = project
    task.save()

    try:
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title='Sites Detail'
        sites = project.sites.all().order_by('identifier')
        if region_ids:
            if isinstance(region_ids, list):
                sites = project.sites.filter(is_active=True, region_id__in=region_ids).order_by('identifier')
            else:
                if region_ids == "0":
                    sites = project.sites.filter(is_active=True, region_id=None).order_by('identifier')
                else:
                    sites = project.sites.filter(is_active=True, region_id=region_ids).order_by('identifier')
        else:
            sites = project.sites.filter(is_active=True).order_by('identifier')

        if type_ids:
            sites = sites.filter(type_id__in=type_ids)

        status, message = siteDetailsGenerator(project, sites, ws)

        if not status:
            raise ValueError(message)

        wb.save(buffer)
        buffer.seek(0)
        xls = buffer.getvalue()
        xls_url = default_storage.save(project.name + '/sites/'+project.name+'-details.xlsx', ContentFile(xls))
        buffer.close()
        task.file.name = xls_url

        task.status = 2
        task.save()

        if sync_to_drive:
            
            if not os.path.exists("media/site-details-report/"):
                os.makedirs("media/site-details-report/")

            temporarylocation="media/site-details-report/site_details_{}.xls".format(project.id)
            with open(temporarylocation,'wb') as out: ## Open temporary file as bytes
                out.write(xls)                ## Read bytes into file

            upload_to_drive(temporarylocation, "{} - Site Information".format(project.id), "Site Information", project, task.user)

            os.remove(temporarylocation)

            task.logs.create(source=task.user, type=32, title="Site details xls sync to Google Drive in project",
                                   recipient=task.user, content_object=task, extra_object=project,
                                   extra_message=" <a href='" +  task.file.url +"'>Xls sites detail report sync to Google Drive in project")
        else:
            task.logs.create(source=task.user, type=32, title="Site details xls generation in project",
                                   recipient=task.user, content_object=task, extra_object=project,
                                   extra_message=" <a href='" +  task.file.url +"'>Xls sites detail report</a> generation in project")


    except DriveException as e:
        print 'Report upload to drive  Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Xls Site Information report upload to Google Drive in Project",
                                       content_object=project, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))
        os.remove(temporarylocation)


    except Exception as e:

        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        print e.__dict__
        task.save()
        task.logs.create(source=task.user, type=432, title="Xls Site Information Report generation in project",
                                   content_object=project, recipient=task.user,
                                   extra_message="@error " + u'{}'.format(e.message))
        buffer.close()


@shared_task(time_limit=600, soft_time_limit=600)
def exportProjectSiteResponses(task_prog_obj_id, project_id, base_url, fs_ids, start_date, end_date, filterRegion, filterSiteTypes):
    time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status = 1
    project=get_object_or_404(Project, pk=project_id)
    task.content_object = project
    task.save()

    try:
        buffer = BytesIO()
        sites = project.sites.filter(is_active=True)
        
        if filterRegion:
            sites = sites.filter(region_id__in=filterRegion)

        if filterSiteTypes:
            sites = sites.filter(type_id__in=filterSiteTypes)
        
        sites = sites.values('id')

        response_sites=[]
        
        split_startdate = start_date.split('-')
        split_enddate = end_date.split('-')

        new_startdate = date(int(split_startdate[0]), int(split_startdate[1]), int(split_startdate[2]))
        end = date(int(split_enddate[0]), int(split_enddate[1]), int(split_enddate[2]))

        new_enddate = end + datetime.timedelta(days=1)

        forms = FieldSightXF.objects.select_related('xf', 'xf__user').filter(pk__in=fs_ids, is_deleted=False)
        wb = Workbook()
        ws_site_details = wb.active
        ws_site_details.title = "Site Details"
        form_id = 0
        form_names=[]

        
        def generate_sheet_name(form_name):
            form_names.append(form_name)
            occurance = form_names.count(form_name)

            sheet_name = form_name[:30]
            
            for ch in ["[", "]", "*", "?", ":", "/"]:
                if ch in sheet_name:
                    sheet_name=sheet_name.replace(ch,"_")

            return sheet_name
        
        for form in forms.iterator():
            form_id += 1
            sheet_name = generate_sheet_name(form.xf.title)
            ws=wb.create_sheet(title=sheet_name)
            
            head_columns = [{'question_name':'No Submission','question_label':'No Submission'}]
            repeat_questions = []
            repeat_answers = []

            ws.append(['Header'])

            
            if filterRegion or filterSiteTypes:
                formresponses = FInstance.objects.select_related('instance', 'site').filter(project_fxf_id=form.id, site_id__in=sites, date__range=[new_startdate, new_enddate])
            else:
                formresponses = FInstance.objects.select_related('instance', 'site').filter(project_fxf_id=form.id, date__range=[new_startdate, new_enddate])

            for formresponse in formresponses.iterator():
                
                if formresponse.site_id and not formresponse.site_id in response_sites:
                    response_sites.append(formresponse.site_id)
                
                questions, answers, r_question_answers = parse_form_response(json.loads(form.xf.json)['children'], formresponse.instance.json, base_url, form.xf.user.username, formresponse.id, True)

                answers['uid'] = formresponse.instance_id
                if formresponse.site_id:
                    answers['identifier'] = formresponse.site.identifier
                    answers['name'] = formresponse.site.name
                    
                else:
                    answers['identifier'] = 'Na'
                    answers['name'] = 'Na'
                
                answers['submitted_by'] = formresponse.submitted_by.email or formresponse.submitted_by.username 
                answers['submitted_on'] = formresponse.instance.date_created
                answers['status'] = form_status_map[formresponse.form_status]
                
                if r_question_answers:
                    repeat_answers.append({'uid':formresponse.instance_id, 'name': answers['name'], 'identifier': answers['identifier'], 'repeated': r_question_answers })

                if len(questions) + 3 > len(head_columns):
                    head_columns = [{'question_name':'uid','question_label':'uid'}, {'question_name':'identifier','question_label':'identifier'}, {'question_name':'name','question_label':'name'}, {'question_name':'submitted_by','question_label':'submitted_by'}, {'question_name':'status','question_label':'status'}, {'question_name':'submitted_on','question_label':'Submitted on'}] + questions  
                row=[]

                for col_num in range(len(head_columns)):
                    row.append(answers.get(head_columns[col_num]['question_name'], ""))    
                ws.append(row)

            for col_num in range(len(head_columns)):
                if isinstance(head_columns[col_num].get('question_label', ""), dict):
                    head_str = head_columns[col_num]['question_label'].get('English (en)', str(head_columns[col_num]['question_label']))
                else:
                    head_str = head_columns[col_num]['question_label']
                ws.cell(row=1, column=col_num+1).value = cleanhtml(head_str)
            
            
            if repeat_answers:
                for group_id, group in repeat_answers[0]['repeated'].items():
                    
                    sheet_name = generate_sheet_name("rep-"+group_id)
                    wr=wb.create_sheet(title=sheet_name)
                    wr.append(['Header'])
                    for repeat in repeat_answers:

                        for answer in repeat['repeated'][group_id]['answers']:
                            row = [repeat['uid'], repeat['identifier'], repeat['name']]
                            for question in group['questions']:
                                row.append(answer.get(question['question_name'], ""))    
                            wr.append(row)
                                

                    wr.cell(row=1, column=1).value = 'uid'
                    wr.cell(row=1, column=2).value = 'Identifier'
                    wr.cell(row=1, column=3).value = 'Name'
                    wr.cell(row=1, column=4).value = 'Submitted by'

                    #for loop needed.
                    for col_num in range(len(group['questions'])):
                        if isinstance(group['questions'][col_num]['question_label'], dict):
                            head_str = group['questions'][col_num]['question_label'].get('English (en)', str(group['questions'][col_num]['question_label']))
                        else:
                            head_str = group['questions'][col_num]['question_label']
                    
                        wr.cell(row=1, column=col_num+3).value = cleanhtml(head_str)
                        
            del formresponses           
            gc.collect()

        if not forms:
            ws = wb.create_sheet(title='No Forms')
        
        elif len(forms) < 2:        
            sites = Site.objects.filter(pk__in=response_sites)
            status, message = siteDetailsGenerator(project, sites, ws_site_details)
            if not status:
                raise ValueError(message)

        del forms
        gc.collect()

        wb.save(buffer)
        buffer.seek(0)
        xls = buffer.getvalue()
        xls_url = default_storage.save(project.name + '/xls/'+project.name+'-submissions.xls', ContentFile(xls))
        buffer.close()

        task.status = 2
        task.file.name = xls_url
        task.save()
        noti = task.logs.create(source=task.user, type=32, title="Xls Report generation in project",
                                   recipient=task.user, content_object=task, extra_object=project,
                                   extra_message=" <a href='"+ task.file.url +"'>Xls report</a> generation in project")
        

    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Report Gen Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Xls Report generation in project",
                                       content_object=project, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))
        buffer.close()

        
@shared_task()
def importSites(task_prog_obj_id, f_project, t_project, meta_attributes, regions, ignore_region):
    time.sleep(2)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.content_object = t_project
    task.status=1
    task.save()
    
    try:
        def filterbyquestion_name(seq, value):
            for el in seq:
                if (not meta_attributes) or (meta_attributes and el.get('question_name') in meta_attributes):
                    if el.get('question_name')==value:
                        return True
            return False
        
        # migrate metas

        if t_project.site_meta_attributes:
            t_metas = t_project.site_meta_attributes
            f_metas = f_project.site_meta_attributes
            
            for f_meta in f_metas:
                # print t_metas
                # print ""
                check = filterbyquestion_name(t_metas, f_meta.get('question_name'))
                if not check:
                    t_metas.append(f_meta)
        region_map = {}      

        t_project_sites = t_project.sites.filter(is_active=True).values_list('identifier', flat=True)

        # migrate regions
        if f_project.cluster_sites and not ignore_region:
            
            t_project_regions = t_project.project_region.filter(is_active=True).values_list('identifier', flat=True)
            t_project.cluster_sites=True
            
            # To handle whole project or a single region migrate
            region_objs = f_project.project_region.filter(id__in=regions, is_active=True)

            for region in region_objs:
                f_region_id = region.id
                if region.identifier in t_project_regions:
                    t_region_id = t_project.project_region.get(identifier=region.identifier, is_active=True).id
                else:
                    region.id=None
                    region.project_id=t_project.id
                    region.save()
                    t_region_id = region.id
                region_map[f_region_id]=t_region_id
        
            t_project.save()

            # getting Sites
        
            sites = f_project.sites.filter(is_active=True, region_id__in=regions)
          
            if 0 in regions:
                unassigned_sites = f_project.sites.filter(is_active=True, region_id=None)
                sites = sites | unassigned_sites

        else:

            sites = f_project.sites.filter(is_active=True)

        
        def get_t_region_id(f_region_id):
            # To get new region id without a query
            if f_region_id is not None and f_region_id in region_map:
                return region_map[f_region_id]
            else:
                return None

        # migrate sites
        for site in sites:
            site.id = None
            site.project_id = t_project.id
            
            if site.identifier in t_project_sites:
                site.identifier = str(site.identifier) + "IFP" + str(f_project.id)
        
            if f_project.cluster_sites and not ignore_region:
                site.region_id = get_t_region_id(site.region_id)
            else:
                site.region_id = None
            
            site.save()

        task.status = 2
        task.save()

        if f_project.cluster_sites and not ignore_region:
            noti = FieldSightLog.objects.create(source=task.user, type=30, title="Bulk Project import sites",
                                       content_object=t_project, recipient=task.user,
                                       extra_object=f_project, extra_message="Project Sites import from " + str(len(regions))+" Regions of ")
        else:
            noti = FieldSightLog.objects.create(source=task.user, type=29, title="Bulk Project import sites",
                                       content_object=t_project, recipient=task.user,
                                       extra_object=f_project)        
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print e.__dict__
        if f_project.cluster_sites and not ignore_region:
            noti = FieldSightLog.objects.create(source=task.user, type=430, title="Bulk Project import sites",
                                       content_object=t_project, recipient=task.user,
                                       extra_object=f_project, extra_message="Project Sites import from "+str(len(regions))+" Regions of ")
        else:
            
            noti = FieldSightLog.objects.create(source=task.user, type=429, title="Bulk Project import sites",
                                       content_object=t_project, recipient=task.user,
                                       extra_object=f_project)           
        

@shared_task()
def multiuserassignproject(task_prog_obj_id, org_id, projects, users, group_id):
    time.sleep(2)
    org = Organization.objects.get(pk=org_id)
    projects_count = len(projects)
    users_count = len(users)
    
    task_id = multiuserassignproject.request.id
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.content_object = org
    task.description = "Assign "+str(users_count)+" people in "+str(projects_count)+" projects."
    task.status=1
    task.save()
    try:
        with transaction.atomic():
            roles_created = 0
            for project_id in projects:
                    project = Project.objects.get(pk=project_id)
                    for user in users:
                        try:
                            role, created = UserRole.objects.get_or_create(user_id=user, project_id=project_id,
                                                                           organization_id=org.id,
                                                                           group_id=group_id, ended_at=None)
                            if created:
                                roles_created += 1
                        
                        except MultipleObjectsReturned:
                            
                            redundant_ids = UserRole.objects.filter(user_id=user, project_id=project_id,
                            organization_id=org.id, group_id=group_id, ended_at=None).order_by('id').values('id')[1:]
                            
                            UserRole.objects.filter(pk__in=redundant_ids).update(ended_at=datetime.datetime.now()) 

                            # description = "{0} was assigned  as Project Manager in {1}".format(
                                # role.user.get_full_name(), role.project)
                            # noti = role.logs.create(source=role.user, type=6, title=description, description=description,
                            #  content_object=role.project, extra_object=self.request.user)
                            # result = {}
                            # result['description'] = description
                            # result['url'] = noti.get_absolute_url()
                            # ChannelGroup("notify-{}".format(role.organization.id)).send({"text": json.dumps(result)})
                            # ChannelGroup("project-{}".format(role.project.id)).send({"text": json.dumps(result)})
                            # ChannelGroup("notify-0").send({"text": json.dumps(result)})
        task.status = 2
        task.save()
        if roles_created == 0:
            noti = FieldSightLog.objects.create(source=task.user, type=23, title="Task Completed.",
                                       content_object=org, recipient=task.user,
                                       extra_message=str(roles_created) + " new Project Manager Roles in " + str(projects_count) + " projects ")
        
        else:
            noti = FieldSightLog.objects.create(source=task.user, type=21, title="Bulk Project User Assign",
                                           content_object=org, organization=org, 
                                           extra_message=str(roles_created) + " new Project Manager Roles in " + str(projects_count) + " projects ")
        
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print e.__dict__
        noti = FieldSightLog.objects.create(source=task.user, type=421, title="Bulk Project User Assign",
                                       content_object=org, recipient=task.user,
                                       extra_message=str(users_count)+" people in "+str(projects_count)+" projects ")

@shared_task()
def multiuserassignsite(task_prog_obj_id, project_id, sites, users, group_id):
    time.sleep(2)
    project = Project.objects.get(pk=project_id)
    group_name = Group.objects.get(pk=group_id).name
    sites_count = len(sites)
    users_count = len(users)

    task_id = multiuserassignsite.request.id
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.content_object = project
    task.description = "Assign "+str(users_count)+" people in "+str(sites_count)+" sites."
    task.status=1
    task.save()
    try:
        with transaction.atomic():
            roles_created = 0            
            for site_id in sites:
                site = Site.objects.get(pk=site_id)
                for user in users:
                    try:
                        role, created = UserRole.objects.get_or_create(user_id=user, site_id=site.id,
                                                                   project__id=project.id, organization__id=site.project.organization_id, group_id=group_id, ended_at=None)
                        if created:
                            roles_created += 1

                    except MultipleObjectsReturned:
                        
                        redundant_ids = UserRole.objects.filter(user_id=user, site_id=site.id, project__id=project.id, organization__id=site.project.organization_id, group_id=group_id, ended_at=None).order_by('id').values('id')[1:]
                        
                        UserRole.objects.filter(pk__in=redundant_ids).update(ended_at=datetime.datetime.now())

                   
                        # description = "{0} was assigned  as {1} in {2}".format(
                        #     role.user.get_full_name(), role.lgroup.name, role.project)
                        # noti_type = 8

                        # if data.get('group') == "Reviewer":
                        #     noti_type =7
                        
                        # noti = role.logs.create(source=role.user, type=noti_type, title=description,
                        #                         description=description, content_type=site, extra_object=self.request.user,
                        #                         site=role.site)
                        # result = {}
                        # result['description'] = description
                        # result['url'] = noti.get_absolute_url()
                        # ChannelGroup("notify-{}".format(role.organization.id)).send({"text": json.dumps(result)})
                        # ChannelGroup("project-{}".format(role.project.id)).send({"text": json.dumps(result)})
                        # ChannelGroup("site-{}".format(role.site.id)).send({"text": json.dumps(result)})
                        # ChannelGroup("notify-0").send({"text": json.dumps(result)})

                        # Device = get_device_model()
                        # if Device.objects.filter(name=role.user.email).exists():
                        #     message = {'notify_type':'Assign Site', 'site':{'name': site.name, 'id': site.id}}
                        #     Device.objects.filter(name=role.user.email).send_message(message)
        task.status = 2
        task.save()
        if roles_created == 0:
            noti = FieldSightLog.objects.create(source=task.user, type=23, title="Task Completed.",
                                       content_object=project, recipient=task.user, 
                                       extra_message="All "+str(users_count) +" users were already assigned as "+ group_name +" in " + str(sites_count) + " selected sites ")
        
        else:

            noti = FieldSightLog.objects.create(source=task.user, type=22, title="Bulk site User Assign",
                                           content_object=project, organization=project.organization, project=project, 
                                           extra_message=str(roles_created) + " new "+ group_name +" Roles in " + str(sites_count) + " sites ")
        
    except Exception as e:
        task.status = 3
        task.description = "ERROR: " + str(e.message) 
        print e.__dict__
        task.save()
        noti = FieldSightLog.objects.create(source=task.user, type=422, title="Bulk Sites User Assign",
                                       content_object=project, recipient=task.user,
                                       extra_message=group_name +" for "+str(users_count)+" people in "+str(sites_count)+" sites ")
        
@shared_task()
def multiuserassignregion(task_prog_obj_id, project_id, regions, users, group_id):
    time.sleep(2)
    project = Project.objects.get(pk=project_id)
    group_name = Group.objects.get(pk=group_id).name
    sites_count = len(regions)
    users_count = len(users)

    task_id = multiuserassignregion.request.id
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.content_object = project
    task.description = "Assign "+str(users_count)+" people in "+str(sites_count)+" regions."
    task.status=1
    task.save()
    try:
        with transaction.atomic():
            roles_created = 0            
            for region_id in regions:
                if region_id == "0":
                    sites = Site.objects.filter(region__isnull=True, project_id=project_id).values('id')
                else: 
                    sites = Site.objects.filter(region_id = region_id, project_id=project_id).values('id')
                for site_id in sites:
                    
                    for user in users:
                        site = Site.objects.filter(pk=site_id['id']).first()
                        if site and site.project_id == project.id:
                            try: 
                                role, created = UserRole.objects.get_or_create(user_id=user, site_id=site_id['id'],
                                                                               project__id=project.id, organization__id=project.organization_id, group_id=group_id, ended_at=None)
                                if created:
                                    roles_created += 1
                            
                            except MultipleObjectsReturned:
                                
                                redundant_ids = UserRole.objects.filter(user_id=user, site_id=site_id['id'],
                                project__id=project.id, organization__id=project.organization_id, group_id=group_id, ended_at=None).order_by('id').values('id')[1:]
                                
                                UserRole.objects.filter(pk__in=redundant_ids).update(ended_at=datetime.datetime.now())
   
                                # description = "{0} was assigned  as {1} in {2}".format(
                                #     role.user.get_full_name(), role.lgroup.name, role.project)
                                # noti_type = 8

                                # if data.get('group') == "Reviewer":
                                #     noti_type =7
                                
                                # noti = role.logs.create(source=role.user, type=noti_type, title=description,
                                #                         description=description, content_type=site, extra_object=self.request.user,
                                #                         site=role.site)
                                # result = {}
                                # result['description'] = description
                                # result['url'] = noti.get_absolute_url()
                                # ChannelGroup("notify-{}".format(role.organization.id)).send({"text": json.dumps(result)})
                                # ChannelGroup("project-{}".format(role.project.id)).send({"text": json.dumps(result)})
                                # ChannelGroup("site-{}".format(role.site.id)).send({"text": json.dumps(result)})
                                # ChannelGroup("notify-0").send({"text": json.dumps(result)})

                                # Device = get_device_model()
                                # if Device.objects.filter(name=role.user.email).exists():
                                #     message = {'notify_type':'Assign Site', 'site':{'name': site.name, 'id': site.id}}
                                #     Device.objects.filter(name=role.user.email).send_message(message)
        task.status = 2
        task.save()
        if roles_created == 0:
            noti = FieldSightLog.objects.create(source=task.user, type=23, title="Task Completed.",
                                       content_object=project, recipient=task.user, 
                                       extra_message="All "+str(users_count) +" users were already assigned as "+ group_name +" in " + str(sites_count) + " selected regions ")
        
        else:

            noti = FieldSightLog.objects.create(source=task.user, type=22, title="Bulk site User Assign",
                                           content_object=project, organization=project.organization, project=project, 
                                           extra_message=str(roles_created) + " new "+ group_name +" Roles in " + str(sites_count) + " regions ")
        
    except Exception as e:
        print 'Bulk role assign Unsuccesfull. ------------------------------------------%s' % e
        task.description = "Assign "+str(users_count)+" people in "+str(sites_count)+" regions. ERROR: " + str(e) 
        task.status = 3
        task.save()
        print e.__dict__
        noti = FieldSightLog.objects.create(source=task.user, type=422, title="Bulk Region User Assign",
                                       content_object=project, recipient=task.user,
                                       extra_message=group_name +" for "+str(users_count)+" people in "+str(sites_count)+" regions ")


@shared_task()
def multi_users_assign_regions(task_prog_obj_id, project_id, regions, users, group_id):
    time.sleep(2)
    project = Project.objects.get(pk=project_id)
    group_name = Group.objects.get(pk=group_id).name
    regions_count = len(regions)
    users_count = len(users)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.content_object = project
    task.description = "Assign " + str(users_count) + " people in " + str(regions_count) + " regions."
    task.status = 1
    task.save()

    # import ipdb;
    # ipdb.set_trace()

    try:
        with transaction.atomic():
            roles_created = 0
            for region_id in regions:
                for user in users:
                    try:
                        role, created = UserRole.objects.get_or_create(user_id=user, project_id=project.id,
                                                                       organization_id=project.organization_id,
                                                                       region_id=region_id,
                                                                       group_id=group_id, ended_at=None)
                        if created:
                            roles_created += 1

                    except MultipleObjectsReturned:

                        redundant_ids = UserRole.objects.filter(user_id=user, project_id=project.id,
                                                                organization_id=project.organization_id,
                                                                group_id=group_id, region_id=region_id, ended_at=None).order_by('id').values('id')[1:]

                        UserRole.objects.filter(pk__in=redundant_ids).update(ended_at=datetime.datetime.now())

        task.status = 2
        task.save()
        if roles_created == 0:
            noti = FieldSightLog.objects.create(source=task.user, type=23, title="Task Completed.",
                                                content_object=project, recipient=task.user,
                                                extra_message="All " + str(
                                                    users_count) + " users were already assigned as " + group_name + " in " + str(
                                                    regions_count) + " selected regions ")

        else:

            noti = FieldSightLog.objects.create(source=task.user, type=22, title="Bulk Region User Assign",
                                                content_object=project, organization=project.organization,
                                                project=project,
                                                extra_message=str(
                                                    roles_created) + " new " + group_name + " Roles in " + str(
                                                    regions_count) + " regions ")

    except Exception as e:
        print 'Bulk role assign Unsuccesfull. ------------------------------------------%s' % e
        task.description = "Assign " + str(users_count) + " people in " + str(regions_count) + " regions. ERROR: " + str(
            e)
        task.status = 3
        task.save()
        print e.__dict__
        noti = FieldSightLog.objects.create(source=task.user, type=422, title="Bulk Region User Assign",
                                            content_object=project, recipient=task.user,
                                            extra_message=group_name + " for " + str(users_count) + " people in " + str(
                                                regions_count) + " regions ")


@shared_task()
def multi_users_assign_to_entire_project(task_prog_obj_id, project_id, regions, users, unassigned_sites):

    time.sleep(2)
    project = Project.objects.get(pk=project_id)
    regions_count = len(regions)
    users_count = len(users)
    unassigned_sites_count = len(unassigned_sites)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.content_object = project
    task.description = "Assign " + str(users_count) + " people in " + str(regions_count) + " regions and " + str(unassigned_sites_count)
    task.status = 1
    task.save()
    region_supervisor = Group.objects.get(name="Region Supervisor")
    site_supervisor = Group.objects.get(name="Site Supervisor")

    try:
        with transaction.atomic():
            roles_created = 0
            for region_id in regions:
                for user in users:
                    try:
                        role, created = UserRole.objects.get_or_create(user_id=user, project_id=project.id,
                                                                       organization_id=project.organization_id,
                                                                       region_id=region_id,
                                                                       group_id=region_supervisor.id, ended_at=None)
                        if created:
                            roles_created += 1

                    except MultipleObjectsReturned:

                        redundant_ids = UserRole.objects.filter(user_id=user, project_id=project.id,
                                                                organization_id=project.organization_id,
                                                                group_id=region_supervisor.id, region_id=region_id, ended_at=None).order_by('id').values('id')[1:]

                        UserRole.objects.filter(pk__in=redundant_ids).update(ended_at=datetime.datetime.now())

        task.status = 2
        task.save()
        if roles_created == 0:
            noti = FieldSightLog.objects.create(source=task.user, type=23, title="Task Completed.",
                                                content_object=project, recipient=task.user,
                                                extra_message="All " + str(
                                                    users_count) + " users were already assigned as " + region_supervisor.name + " in " + str(
                                                    regions_count) + " selected regions ")

        else:

            noti = FieldSightLog.objects.create(source=task.user, type=22, title="Bulk Region User Assign",
                                                content_object=project, organization=project.organization,
                                                project=project,
                                                extra_message=str(
                                                    roles_created) + " new " + region_supervisor.name + " Roles in " + str(
                                                    regions_count) + " regions ")

    except Exception as e:
        print 'Bulk role assign Unsuccesfull. ------------------------------------------%s' % e
        task.description = "Assign " + str(users_count) + " people in " + str(regions_count) + " regions. ERROR: " + str(
            e)
        task.status = 3
        task.save()
        print e.__dict__
        noti = FieldSightLog.objects.create(source=task.user, type=422, title="Bulk Region User Assign",
                                            content_object=project, recipient=task.user,
                                            extra_message=region_supervisor.name + " for " + str(users_count) + " people in " + str(
                                                regions_count) + " regions ")


@shared_task(time_limit=18000, soft_time_limit=18000)
def auto_generate_stage_status_report():
    projects = Project.objects.filter(active=True)
    for project in projects:
        if Site.objects.filter(project_id=project.id).count() < 2000:
            continue
        else:    
            try:
                data = []
                ss_index = {}
                stages_rows = []
                head_row = ["Site ID", "Name", "Region ID", "Latitude", "longitude", "Status"]
                
                stages = project.stages.filter(stage__isnull=True)
                for stage in stages:
                    sub_stages = stage.parent.all()
                    if len(sub_stages):
                        head_row.append("Stage :"+stage.name)
                        stages_rows.append("Stage :"+stage.name)

                        for ss in sub_stages:
                            head_row.append("Sub Stage :"+ss.name)
                            ss_index.update({head_row.index("Sub Stage :"+ss.name): ss.id})
                head_row.extend(["Site Visits", "Submission Count", "Flagged Submission", "Rejected Submission"])
                data.append(head_row)
                total_cols = len(head_row) - 6 # for non stages
                for site in project.sites.filter(is_active=True, is_survey=False):
                    flagged_count = 0 
                    rejected_count = 0
                    submission_count = 0

                    if site.region:
                        site_row = [site.identifier, site.name, site.region.identifier, site.latitude, site.longitude, site.site_status]
                    else:
                        site_row = [site.identifier, site.name, site.region_id, site.latitude, site.longitude, site.site_status]

                    site_row.extend([None]*total_cols)
                    for k, v in ss_index.items():
                        if Stage.objects.filter(id=v).count() == 1:
                            site_sub_stage = Stage.objects.get(id=v)
                            site_row[k] = site_sub_stage.site_submission_count(v, site.id)
                            submission_count += site_row[k]
                            flagged_count += site_sub_stage.flagged_submission_count(v, site.id)
                            rejected_count += site_sub_stage.rejected_submission_count(v, site.id)
                        else:
                            site_row[k] = 0



                    site_visits = settings.MONGO_DB.instances.aggregate([{"$match":{"fs_site": str(site.id)}},  { "$group" : { 
                          "_id" :  
                            { "$substr": [ "$start", 0, 10 ] }
                          
                       }
                     }])['result']

                    site_row[-1] = rejected_count
                    site_row[-2] = flagged_count
                    site_row[-3] = submission_count
                    site_row[-4] = len(site_visits) 

                    data.append(site_row)

                p.save_as(array=data, dest_file_name="media/stage-report/{}_stage_data.xls".format(project.id))
                xl_data = open("media/stage-report/{}_stage_data.xls".format(project.id), "rb")
                
                #Its only quick fix for now, save it in aws bucket whenever possible.

                project.progress_report = xl_data.name
                project.save()
                
            except Exception as e:
                print 'Report Gen Unsuccesfull. %s' % e
                print e.__dict__
                

def sendNotification(notification, recipient):
    result={}
    result['id']= noti.id,
    result['source_uid']= source_user.id,
    result['source_name']= source_user.username,
    result['source_img']= source_user.user_profile.profile_picture.url,
    result['get_source_url']= noti.get_source_url(),
    result['get_event_name']= project.name,
    result['get_event_url']= noti.get_event_url(),
    result['get_extraobj_name']= None,
    result['get_extraobj_url']= None,
    result['get_absolute_url']= noti.get_absolute_url(),
    result['type']= 412,
    result['date']= str(noti.date),
    result['extra_message']= str(count) + " Sites @error " + u'{}'.format(e.message),
    result['seen_by']= [],
    ChannelGroup("notif-user-{}".format(recipient.id)).send({"text": json.dumps(result)})


@shared_task(time_limit=120, soft_time_limit=120)
def exportProjectstatistics(task_prog_obj_id, project_id, reportType, start_date, end_date):
    time.sleep(3)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status = 1
    project=get_object_or_404(Project, pk=project_id)
    task.content_object = project
    task.save()

    try:
        buffer = BytesIO()
        sites = project.sites.filter(is_active=True)
        data = []
        index = {}
        split_startdate = start_date.split('-')
        split_enddate = end_date.split('-')

        new_startdate = date(int(split_startdate[0]), int(split_startdate[1]), int(split_startdate[2]))
        end = date(int(split_enddate[0]), int(split_enddate[1]), int(split_enddate[2]))

        new_enddate = end + datetime.timedelta(days=1)
        query = {}
        query['pending'] = Sum(
            Case(
                When(fieldsight_instance__form_status=0, then=1),
                default=0, output_field=IntegerField()
            ))

        query['approved'] = Sum(
            Case(
                When(fieldsight_instance__form_status=3, then=1),
                default=0, output_field=IntegerField()
            )
        )
        query['flagged'] = Sum(
            Case(
                When(fieldsight_instance__form_status=2, then=1),
                default=0, output_field=IntegerField()
            ))

        query['rejected'] = Sum(
            Case(
                When(fieldsight_instance__form_status=1, then=1),
                default=0, output_field=IntegerField()
            )
        )

        review_query = {}
        review_query['re_flagged'] = Sum(
            Case(
                When(new_status=2, then=1),
                default=0, output_field=IntegerField()
            ))        

        review_query['re_approved'] = Sum(
            Case(
                When(new_status=3, then=1),
                default=0, output_field=IntegerField()
            ))   

        review_query['re_rejected'] = Sum(
            Case(
                When(new_status=1, then=1),
                default=0, output_field=IntegerField()
            ))   

        review_query['resolved'] = Sum(
            Case(
                When(old_status__in=[1,2], new_status=3, then=1),
                default=0, output_field=IntegerField()
            ))

        if reportType == "Monthly":
            data.insert(0, ["Date", "Month", "Site Visits", "Submissions", "Active Users", "Approved Submissions", "Pending Submissions", "Rejected Submissions", "Flagged Submissions", "Submission Reviews",  "Resolved Submissions", "Approved Reviews", "Rejected Reviews", "Flagged Reviews"])
            i=1
            for month in rrule(MONTHLY, dtstart=new_startdate, until=end):
                str_month = month.strftime("%Y-%m")
                data.insert(i, [str_month, month.strftime("%B"), 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,])
                index[str_month] = i
                i += 1

            site_visits = settings.MONGO_DB.instances.aggregate([{"$match":{"fs_project": {"$in":[project_id, int(project_id)]}, "start": 
              { '$gte' : new_startdate.isoformat(), '$lte' : new_enddate.isoformat() }}},  { "$group" : { 
                                "_id" :  { 
                                  
                                  "fs_site": "$fs_site",
                                  "date": { "$substr": [ "$start", 0, 10 ] }
                                },
                             }
                           }, { "$group": { "_id": "$_id.date", "visits": { '$sum': 1}
                           }},
                           {"$group": {"_id": { "$substr": [ "$_id", 0, 7 ] }, "total_sum": {'$sum': '$visits'}}}
                           ])['result']

            for visit in site_visits:
                if visit['_id'] != "":
                    data[index[visit['_id']]][2] = int(visit['total_sum'])

            truncate_date = connection.ops.date_trunc_sql('month', 'logger_instance.date_created')
            forms=Instance.objects.filter(fieldsight_instance__project_id=project_id, date_created__gte=new_startdate,  date_created__lte=new_enddate).extra({'date_created':truncate_date})
            
            forms_stats=forms.values('date_created').annotate(dcount=Count('date_created'), **query)

            for month_stat in forms_stats:
                if month_stat['date_created'].strftime("%Y-%m") in index:
                    data[index[month_stat['date_created'].strftime("%Y-%m")]][3] = int(month_stat['dcount'])
                    data[index[month_stat['date_created'].strftime("%Y-%m")]][5] = int(month_stat['approved'])
                    data[index[month_stat['date_created'].strftime("%Y-%m")]][6] = int(month_stat['pending'])
                    data[index[month_stat['date_created'].strftime("%Y-%m")]][7] = int(month_stat['rejected'])
                    data[index[month_stat['date_created'].strftime("%Y-%m")]][8] = int(month_stat['flagged'])
                
            truncate_date = connection.ops.date_trunc_sql('month', 'fsforms_instancestatuschanged.date')
            status_changed=InstanceStatusChanged.objects.filter(finstance__project_id=project_id, date__range=[new_startdate, new_enddate]).extra({'date':truncate_date})
            status_months=status_changed.values('date').annotate(dcount=Count('date'), **review_query)
            for status_month in status_months:
                if status_month['date'].strftime("%Y-%m") in index:
                
                    data[index[status_month['date'].strftime("%Y-%m")]][9] = int(status_month['dcount'])
                    data[index[status_month['date'].strftime("%Y-%m")]][10] = int(status_month['resolved'])
                    data[index[status_month['date'].strftime("%Y-%m")]][11] = int(status_month['re_approved'])
                    data[index[status_month['date'].strftime("%Y-%m")]][12] = int(status_month['re_rejected'])
                    data[index[status_month['date'].strftime("%Y-%m")]][13] = int(status_month['re_flagged'])
                
            truncate_date = connection.ops.date_trunc_sql('month', 'logger_instance.date_created')
            forms=Instance.objects.filter(fieldsight_instance__project_id=project_id, date_created__range=[new_startdate, new_enddate]).extra({'date_created':truncate_date})
            forms_stats=forms.values('date_created').annotate(dcount=Count('user_id', distinct=True))

            for month_stat in forms_stats:
                try:
                    data[index[month_stat['date_created'].strftime("%Y-%m")]][4] = int(month_stat['dcount'])

                except:
                    pass

        if reportType in ["Daily", "Weekly"]:
            data.insert(0, ["Date", "Day", "Site Visits", "Submissions", "Active Users", "Approved Submissions", "Pending Submissions", "Rejected Submissions", "Flagged Submissions", "Submission Reviews",  "Resolved Submissions", "Approved Reviews", "Rejected Reviews", "Flagged Reviews"])
            i=1
            for day in rrule(DAILY, dtstart=new_startdate, until=end):
                str_day = day.strftime("%Y-%m-%d")
                data.insert(i, [str_day, day.strftime("%A"), 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, ])
                index[str_day] = i
                i += 1

            site_visits = settings.MONGO_DB.instances.aggregate([{"$match":{"fs_project": {"$in":[project_id, int(project_id)]}, "start": 
              { '$gte' : new_startdate.isoformat(), '$lte' : new_enddate.isoformat() }}},  { "$group" : { 
                                "_id" :  { 
                                  
                                  "fs_site": "$fs_site",
                                  "date": { "$substr": [ "$start", 0, 10 ] }
                                },
                             }
                           }, { "$group": { "_id": "$_id.date", "visits": { '$sum': 1}
                           }},
                           {"$group": {"_id": { "$substr": [ "$_id", 0, 10 ] }, "total_sum": {'$sum': '$visits'}}}
                           ])['result']

            for visit in site_visits:
                if visit['_id'] != "":
                    data[index[visit['_id']]][2] = int(visit['total_sum'])

            truncate_date = connection.ops.date_trunc_sql('day', 'date_created')
            forms=Instance.objects.filter(fieldsight_instance__project_id=project_id, date_created__range=[new_startdate, new_enddate]).extra({'date_created':truncate_date})
            forms_stats=forms.values('date_created').annotate(dcount=Count('date_created'), **query)
        
            for day_stat in forms_stats:
                if day_stat['date_created'].strftime("%Y-%m-%d") in index:
                    data[index[day_stat['date_created'].strftime("%Y-%m-%d")]][3] = int(day_stat['dcount'])
                    data[index[day_stat['date_created'].strftime("%Y-%m-%d")]][5] = int(day_stat['approved'])
                    data[index[day_stat['date_created'].strftime("%Y-%m-%d")]][6] = int(day_stat['pending'])
                    data[index[day_stat['date_created'].strftime("%Y-%m-%d")]][7] = int(day_stat['rejected'])
                    data[index[day_stat['date_created'].strftime("%Y-%m-%d")]][8] = int(day_stat['flagged'])

             
            truncate_date = connection.ops.date_trunc_sql('day', 'fsforms_instancestatuschanged.date')
            status_changed=InstanceStatusChanged.objects.filter(finstance__project_id=project_id, date__range=[new_startdate, new_enddate]).extra({'date':truncate_date})
            status_days=status_changed.values('date').annotate(dcount=Count('date'), **review_query)

            for status_day in status_days:
                if status_day['date'].strftime("%Y-%m-%d") in index:
                    data[index[status_day['date'].strftime("%Y-%m-%d")]][9] = int(status_day['dcount'])
                    data[index[status_day['date'].strftime("%Y-%m-%d")]][10] = int(status_day['resolved'])
                    data[index[status_day['date'].strftime("%Y-%m-%d")]][11] = int(status_day['re_approved'])
                    data[index[status_day['date'].strftime("%Y-%m-%d")]][12] = int(status_day['re_rejected'])
                    data[index[status_day['date'].strftime("%Y-%m-%d")]][13] = int(status_day['re_flagged'])


            truncate_date = connection.ops.date_trunc_sql('day', 'date_created')
            forms=Instance.objects.filter(fieldsight_instance__project_id=project_id, date_created__range=[new_startdate, new_enddate]).extra({'date_created':truncate_date})
            forms_stats=forms.values('date_created').annotate(dcount=Count('user_id', distinct=True))


            for month_stat in forms_stats:
                if month_stat['date_created'].strftime("%Y-%m-%d") in index:
                    data[index[month_stat['date_created'].strftime("%Y-%m-%d")]][4] = int(month_stat['dcount'])

        wb = Workbook()
        ws = wb.active
        ws.title = "Site Status"

        if reportType == "Weekly":
            weekly_data = [["Week No.", "Week Start", "Week End", "Site Visits", "Submissions", "Active Users", "Approved Submissions", "Pending Submissions", "Rejected Submissions", "Flagged Submissions", "Submission Reviews",  "Resolved Submissions", "Approved Reviews", "Rejected Reviews", "Flagged Reviews"]]
            weekcount = 0
            for value in data[1:]:
                day = datetime.datetime.strptime(value[0], "%Y-%m-%d").weekday() + 1
                # Since start day is Monday And in Nepa we Calculate from Saturday for now.
                if day == 7 or weekcount == 0:
                    weekcount += 1
                    weekly_data.insert(weekcount, ["Week "+ str(weekcount), "", "", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,])

                    weekly_data[weekcount][1] = value[0]
                weekly_data[weekcount][2] = value[0]
                weekly_data[weekcount][3] += value[2]
                weekly_data[weekcount][4] += value[3]
                weekly_data[weekcount][5] += value[4]
                weekly_data[weekcount][6] += value[5]
                weekly_data[weekcount][7] += value[6]
                weekly_data[weekcount][8] += value[7]
                weekly_data[weekcount][9] += value[8]
                weekly_data[weekcount][10] += value[9]
                weekly_data[weekcount][11] += value[10]
                weekly_data[weekcount][12] += value[11]
                weekly_data[weekcount][13] += value[12]
                weekly_data[weekcount][14] += value[13]

                 

            for value in weekly_data:
                ws.append(value)

        else:
            for value in data:
                ws.append(value)

        wb.save(buffer)
        buffer.seek(0)
        xls = buffer.getvalue()
        xls_url = default_storage.save(project.name + '/xls/'+project.name+'-statistics.xls', ContentFile(xls))
        buffer.close()

        task.status = 2
        task.file.name = xls_url
        task.save()
        noti = task.logs.create(source=task.user, type=32, title="Xls Project stastics Report generation in project",
                                 recipient=task.user, content_object=task, extra_object=project,
                                 extra_message=" <a href='"+ task.file.url +"'>Xls project statistics report</a> generation in project")

    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Report Gen Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Xls project statistics report generation in project",
                                       content_object=project, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))
        buffer.close()


@shared_task(time_limit=120, soft_time_limit=120)
def exportLogs(task_prog_obj_id, pk, reportType, start_date, end_date):
    time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status = 1
    if reportType == "Project":
        obj=get_object_or_404(Project, pk=pk)
    else:
        obj=get_object_or_404(Site, pk=pk)

    task.content_object = obj
    task.save()

    try: 
        buffer = BytesIO()
        data = []
        index = {}
        split_startdate = start_date.split('-')
        split_enddate = end_date.split('-')

        new_startdate = date(int(split_startdate[0]), int(split_startdate[1]), int(split_startdate[2]))
        end = date(int(split_enddate[0]), int(split_enddate[1]), int(split_enddate[2]))

        new_enddate = end + datetime.timedelta(days=1)
        queryset = FieldSightLog.objects.select_related('source__user_profile').filter(recipient=None, date__range=[new_startdate, new_enddate]).exclude(type__in=[23, 29, 30, 32, 35])
        
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Day and Time", "User", "Log"])
        try:
            offset_time = task.user.user_profile.timezone.offset_time if task.user.user_profile.timezone.offset_time else "UTC +05:45"
        except:
            offset_time = "UTC +05:45"

        operator = offset_time[4]
        time_offset = offset_time[5:]
        hour_offset = time_offset.split(':')[0]
        minute_offset = time_offset.split(':')[1]

        if reportType == "Project":
            ws.title = "Project Logs"
            logs = queryset.filter(Q(project_id=pk) | (Q(content_type=ContentType.objects.get(app_label="fieldsight", model="project")) & Q(object_id=pk)))

        else:
            ws.title = "Site Logs"
            site = Site.objects.get(pk=pk)
            content_site = ContentType.objects.get(app_label="fieldsight", model="site")
            project = site.project
            query = Q(site_id=pk) | (Q(content_type=content_site) & Q(object_id=pk)) | (Q(extra_content_type=content_site) & Q(extra_object_id=pk))
            meta_dict = {}
            for meta in project.site_meta_attributes:
                if meta['question_type'] == "Link" and meta['question_name'] in site.site_meta_attributes_ans:
                    meta_site_id = Site.objects.filter(identifier=site.site_meta_attributes_ans[meta['question_name']], project_id=meta['project_id'])
                    if meta_site_id:
                        selected_metas = [sub_meta['question_name'] for sub_meta in meta['metas'][str(meta['project_id'])]]
                        meta_dict[meta_site_id[0].id] = selected_metas

            for key, value in meta_dict.items():
                for item in value:
                    query |= (Q(type=15) & Q(content_type=content_site) & Q(object_id=key) & Q(extra_json__contains='"'+item +'":'))

            logs = queryset.filter(query)            
        
        local_log_types = log_types

        for log in logs:
            if operator == '+':
                day_time = log.date + datetime.timedelta(hours=int(hour_offset), minutes=int(minute_offset))
            else:
                day_time = log.date - datetime.timedelta(hours=int(hour_offset), minutes=int(minute_offset))

            day_time = day_time.strftime('%A, %-I:%-M %p')

            if log.type == 15:
                log_text, sub_log_text  = local_log_types[str(log.type)](log)    
                row_data = [log.date, day_time, log.source.first_name + ' ' + log.source.last_name, log_text]
                
                ws.append(row_data)
                if sub_log_text:
                    for sub_log in sub_log_text:
                        ws.append(sub_log)
            
            else:                
                log_text = local_log_types[str(log.type)](log)
                ws.append([log.date, day_time, log.source.first_name + ' ' + log.source.last_name, log_text])

        wb.save(buffer)
        buffer.seek(0)
        xls = buffer.getvalue()
        base_uri = obj.name if reportType == "Project" else obj.project.name + '/' + obj.name
        xls_url = default_storage.save(base_uri + '/xls/'+obj.name+'-logs.xls', ContentFile(xls))
        buffer.close()

        task.status = 2
        task.file.name = xls_url
        task.save()
        noti = task.logs.create(source=task.user, type=32, title="Xls "+ reportType +" Logs Report generation",
                                 recipient=task.user, content_object=task, extra_object=obj,
                                 extra_message=" <a href='"+ task.file.url +"'>Xls "+ reportType +" statistics report</a> generation in ")
# user = User.objects.get(username="fsadmin")
# exportLogs( 2143, user , 137, "Project", "2018-08-11", "2018-12-12")
    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Report Gen Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Xls "+ reportType +" logs report generation in ",
                                       content_object=obj, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))
        buffer.close()


@shared_task(time_limit=120, soft_time_limit=120)
def exportProjectUserstatistics(task_prog_obj_id, project_id, start_date, end_date):
    # time.sleep(5)
    task = CeleryTaskProgress.objects.get(pk=task_prog_obj_id)
    task.status = 1
    project=get_object_or_404(Project, pk=project_id)
    task.content_object = project
    task.save()

    try:  
        buffer = BytesIO()
        sites = project.sites.filter(is_active=True)
        index = {}
        split_startdate = start_date.split('-')
        split_enddate = end_date.split('-')

        new_startdate = date(int(split_startdate[0]), int(split_startdate[1]), int(split_startdate[2]))
        end = date(int(split_enddate[0]), int(split_enddate[1]), int(split_enddate[2]))

        new_enddate = end + datetime.timedelta(days=1)

       
        headers = ["UserName", "Full name", "Email", "Submssions made", "Sites Visited", "Days worked", "Submissions last month", "Submissions last week", "submissions Today", "Approved Submissions", "Pending Submissions", "Flagged Submissions", "Rejected Submissions", "Reviewed Submissions", "Resolved Submissions", "Approved Reviews", "Flagged Reviews", "Rejected Reviews"]

        site_visits = settings.MONGO_DB.instances.aggregate(
            [
                {
                    "$match":{
                        "fs_project": {
                            "$in":[project_id, int(project_id)]
                        },
                        "start": { 
                            '$gte' : new_startdate.isoformat(),
                            '$lte' : new_enddate.isoformat() 
                        },
                        "fs_project": {'$in' : [str(project_id), int(project_id)]}
                    }
                },
                { 
                    "$group" : { 
                        "_id" :  { 
                            "user": "$_submitted_by",
                            "fs_site": "$fs_site",
                            "date": { 
                                "$substr": [ "$start", 0, 10 ]
                            }
                        },
                            "submissions": {'$sum':1}
                    }
                },
                {
                    "$group": {
                        "_id": {
                            "_user":"$_id.user",
                            "_fs_site": "$_id.fs_site"
                        },
                        "submissions": {'$sum': '$submissions'},
                        "visits": { '$sum': 1}
                    }
                },
                {
                    "$group": {
                        "_id": "$_id._user",
                        "submissions": {'$sum': '$submissions'},
                        "sites_visited": {'$sum': 1}
                    }
                }
            ]
        )['result']

        all_days_worked = settings.MONGO_DB.instances.aggregate(
            [
                {
                    "$match":{
                        "fs_project": {
                            "$in":[project_id, int(project_id)]
                        },
                        "start": { 
                            '$gte' : new_startdate.isoformat(),
                            '$lte' : new_enddate.isoformat() 
                        },
                        "fs_project": {'$in' : [str(project_id), int(project_id)]}
                    }
                },
                { 
                    "$group" : { 
                        "_id" :  { 
                            "user": "$_submitted_by",
                            "date": { 
                                "$substr": [ "$start", 0, 10 ]
                            }
                        },
                            "submissions": {'$sum':1}
                    }
                },
                {
                    "$group": {
                        "_id": "$_id.user",
                        "days_worked": { '$sum': 1}
                    }
                }
            ]
        )['result']

        user_stats = {}

        for visit in site_visits:
            visit['total_worked_days'] = 0
            user_stats[visit['_id']] = visit

        for days_worked in all_days_worked:
            user_stats[days_worked['_id']]['total_worked_days'] = days_worked['days_worked']

        query={}
        last_month = new_enddate - datetime.timedelta(days=30)
        query['all_submissions'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[new_startdate, new_enddate], supervisor__project_id=project_id, then=1),
                default=0, output_field=IntegerField()
            ))
        query['monthly'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[last_month, new_enddate], supervisor__project_id=project_id, then=1),
                default=0, output_field=IntegerField()
            ))
        last_week = new_enddate - datetime.timedelta(days=7)
        query['weekly'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[last_week, new_enddate], supervisor__project_id=project_id,then=1),
                default=0, output_field=IntegerField()
            ))

        query['daily'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[end, new_enddate], supervisor__project_id=project_id, then=1),
                default=0, output_field=IntegerField()
            ))

        query['pending'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[new_startdate, new_enddate],supervisor__form_status=0, supervisor__project_id=project_id, then=1),
                default=0, output_field=IntegerField()
            ))

        query['rejected'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[new_startdate, new_enddate],supervisor__form_status=1, supervisor__project_id=project_id, then=1),
                default=0, output_field=IntegerField()
            ))

        query['flagged'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[new_startdate, new_enddate],supervisor__form_status=2, supervisor__project_id=project_id, then=1),
                default=0, output_field=IntegerField()
            ))

        query['approved'] = Sum(
            Case(
                When(supervisor__instance__date_created__range=[new_startdate, new_enddate],supervisor__form_status=3, supervisor__project_id=project_id, then=1),
                default=0, output_field=IntegerField()
            ))    

        review_query = {}
        review_query['re_approved'] = Sum(
            Case(
                When(submission_comments__date__range=[new_startdate, new_enddate], submission_comments__finstance__project_id=project_id, submission_comments__new_status=3, then=1),
                default=0, output_field=IntegerField()
            ))

        review_query['re_rejected'] = Sum(
            Case(
                When(submission_comments__date__range=[new_startdate, new_enddate], submission_comments__finstance__project_id=project_id, submission_comments__new_status=1, then=1),
                default=0, output_field=IntegerField()
            ))        

        review_query['re_flagged'] = Sum(
            Case(
                When(submission_comments__date__range=[new_startdate, new_enddate], submission_comments__finstance__project_id=project_id, submission_comments__new_status=2, then=1),
                default=0, output_field=IntegerField()
            ))        

        review_query['resolved'] = Sum(
            Case(
                When(submission_comments__date__range=[new_startdate, new_enddate], submission_comments__finstance__project_id=project_id, submission_comments__old_status__in=[1,2], submission_comments__new_status=3, then=1),
                default=0, output_field=IntegerField()
            ))        

        dumb_visits = {
            "total_worked_days": 0,
            "submissions": 0,
            "sites_visited": 0
        }

        users=User.objects.filter(user_roles__project_id=project_id, user_roles__group_id__in=[2, 3, 4, 9]).distinct('id').values('id')
        activity_dict = {}
        for user in User.objects.filter(pk__in=users).annotate(**query):
            activity_dict[str(user.id)] = [user.username, user.get_full_name(), user.email, user.all_submissions, user_stats.get(user.username, dumb_visits)['sites_visited'], user_stats.get(user.username, dumb_visits)['total_worked_days'], user.monthly, user.weekly, user.daily, user.approved, user.pending, user.flagged, user.rejected, 0, 0, 0, 0, 0]

        for user in User.objects.filter(pk__in=users).annotate(**review_query):
            activity_dict[str(user.id)][13:18] = [user.re_approved + user.re_rejected + user.re_flagged, user.resolved, user.re_approved, user.re_flagged, user.re_rejected]

        wb = Workbook()
        ws = wb.active
        sheet_name = "Report "+start_date+"_"+end_date
        sheet_name = sheet_name[:30]
        for ch in ["[", "]", "*", "?", ":", "/"]:
            if ch in sheet_name:
                sheet_name=sheet_name.replace(ch,"_")
        ws.title=sheet_name
        ws.append(headers)
        for key, row in activity_dict.items():
            ws.append(row)
        wb.save(buffer)
        buffer.seek(0)
        xls = buffer.getvalue()
        xls_url = default_storage.save(project.name + '/xls/'+project.name+'-User statistics.xls', ContentFile(xls))
        buffer.close()

        task.status = 2
        task.file.name = xls_url
        task.save()
        noti = task.logs.create(source=task.user, type=32, title="Xls Project User stastics Report generation in project",
                                 recipient=task.user, content_object=task, extra_object=project,
                                 extra_message=" <a href='"+ task.file.url +"'>Xls project user statistics report</a> generation in project")

    except Exception as e:
        task.description = "ERROR: " + str(e.message) 
        task.status = 3
        task.save()
        print 'Report Gen Unsuccesfull. %s' % e
        print e.__dict__
        noti = task.logs.create(source=task.user, type=432, title="Xls project user statistics report generation in project",
                                       content_object=project, recipient=task.user,
                                       extra_message="@error " + u'{}'.format(e.message))
        buffer.close()
#
# @shared_task(max_retries=5)
# def auto_create_default_project_site(user, organization_id):
#     project_type_id = ProjectType.objects.first().id
#     project = Project.objects.create(name="Demo Project", organization_id=organization_id, type_id=project_type_id)
#     print('project createed')
#     Site.objects.create(name="Demo Site", project=project)
#     print('site createed')
#     token = user.auth_token.key
#     clone_form.delay(user, token, project)

@shared_task(time_limit=120, soft_time_limit=120)
def email_after_signup(user_id, to_email):
    time.sleep(10)
    user = User.objects.get(id=user_id)
    mail_subject = 'Activate your account.'
    message = render_to_string('users/acc_active_email.html', {
        'user': user,
        'domain': settings.SITE_URL,
        'uid': urlsafe_base64_encode(force_bytes(user_id)),
        'token': account_activation_token.make_token(user),
    })

    email = EmailMessage(
        mail_subject, message, to=[to_email]
    )
    email.content_subtype = "html"
    email.send()


@shared_task(time_limit=120, soft_time_limit=120)
def email_after_subscribed_plan(user_id):
    free_package = Package.objects.get(plan=0)
    user = User.objects.get(id=user_id)
    mail_subject = 'Thank you'
    message = render_to_string('subscriptions/subscribed_email.html', {
        'user': user.username,
        'plan': free_package,
        'domain': settings.SITE_URL,
    })
    to_email = user.email
    email = EmailMessage(
        mail_subject, message, to=[to_email]
    )
    email.content_subtype = "html"
    email.send()


def warning_emails(subscriber, plan_name, total_submissions, extra_submissions_charge, period_type, usage_rates, renewal_date, email):

    mail_subject = 'Warning'
    message = render_to_string('subscriptions/warning_email.html', {
        'user': subscriber.stripe_customer.user.first_name,
        'plan_name': plan_name,
        'total_submissions': total_submissions,
        'extra_submissions_charge': extra_submissions_charge,
        'usage_rates': usage_rates,
        'period_type': period_type,
        'renewal_date': renewal_date.strftime('%B-%d-%Y')
    })
    to_email = email
    email = EmailMessage(
        mail_subject, message, to=[to_email]
    )
    email.content_subtype = "html"
    email.send()


@shared_task()
def warning_overage_emails(subscriber, plan_name, total_submissions, extra_submissions_charge, renewal_date, email):

    mail_subject = 'Warning'
    message = render_to_string('subscriptions/warning_overage_email.html', {
        'user': subscriber.stripe_customer.user.first_name,
        'plan_name': plan_name,
        'total_submissions': total_submissions,
        'extra_submissions_charge': extra_submissions_charge,
        'renewal_date': renewal_date.strftime('%B-%d-%Y')

    })
    to_email = email
    email = EmailMessage(
        mail_subject, message, to=[to_email]
    )
    email.content_subtype = "html"
    email.send()



@shared_task()
def warning_overage_emails_monthly(subscriber, plan_name, total_submissions, extra_submissions_charge, renewal_date, email):

    mail_subject = 'Warning'
    message = render_to_string('subscriptions/warning_overage_email.html', {
        'user': subscriber.stripe_customer.user.first_name,
        'plan_name': plan_name,
        'total_submissions': total_submissions,
        'extra_submissions_charge': extra_submissions_charge,
        'renewal_date': renewal_date.strftime('%B-%d-%Y')

    })
    to_email = email
    email = EmailMessage(
        mail_subject, message, to=[to_email]
    )
    email.content_subtype = "html"
    email.send()

    length_of_day = renewal_date - datetime.datetime.today()

    if length_of_day.days > 30:
        """
            Send email after 30 days if upcoming renewal date is more than 30 days.
        """

    after_one_month = datetime.datetime.today() + datetime.timedelta(days=30)
    warning_overage_emails_monthly.apply_async(
        args=[subscriber, plan_name, total_submissions, extra_submissions_charge, renewal_date, email],
        eta=after_one_month)


@shared_task()
def check_usage_rates():
    print(".......Checking Usage rates.......")
    subscriptions = Subscription.objects.all().select_related('stripe_customer', 'organization', 'package')

    for subscriber in subscriptions:
        total_submissions = subscriber.package.submissions
        usage_submission = subscriber.organization.get_total_submissions()
        usage_rates = (usage_submission/total_submissions)*100
        email = subscriber.stripe_customer.user.email
        plan_name = subscriber.package.get_plan_display()
        extra_submissions_charge = subscriber.package.extra_submissions_charge
        period_type = subscriber.package.get_period_type_display()
        started_data = subscriber.initiated_on

        if period_type == "Month":
            today_month = datetime.datetime.now().date().month
            today_year = datetime.datetime.now().date().year
            renewal_date = started_data + dateutil.relativedelta.relativedelta(months=1)
            track_periodic_warning_email_obj = TrackPeriodicWarningEmail.objects.filter(subscriber=subscriber,
                                                                                        is_email_send=True,
                                                                                        date__month=today_month,
                                                                                        date__year=today_year
                                                                                        ).exists()

        elif period_type == "Year":
            today = datetime.datetime.now().date().year
            renewal_date = started_data + dateutil.relativedelta.relativedelta(months=12)
            track_periodic_warning_email_obj = TrackPeriodicWarningEmail.objects.filter(subscriber=subscriber,
                                                                                        is_email_send=True,
                                                                                        date__year=today).exists()

        if 75 <= usage_rates < 76:
            warning_emails(subscriber, plan_name, total_submissions, extra_submissions_charge, period_type, usage_rates, renewal_date, email)

        elif 90 <= usage_rates < 91:
            warning_emails(subscriber, plan_name, total_submissions, extra_submissions_charge, period_type, usage_rates, renewal_date, email)

        elif 95 <= usage_rates < 96:
            warning_emails(subscriber, plan_name, total_submissions, extra_submissions_charge, period_type, usage_rates, renewal_date, email)

        elif usage_submission >= total_submissions:

            if not track_periodic_warning_email_obj:
                """
                    Send Warning E-Mails when total usage reached and overage charges begin, and then at 1 day, 3 days, 1 week,
                    and then monthly (for annual plans)
                """
                TrackPeriodicWarningEmail.objects.create(subscriber=subscriber, is_email_send=True, date=datetime.datetime.now().date())
                warning_overage_emails.delay(subscriber, plan_name, total_submissions, extra_submissions_charge,
                                             renewal_date, email)

                length_of_day = renewal_date - datetime.datetime.today()

                if length_of_day.days > 3:
                    """
                        Send email after 3 days if upcoming renewal date is more than 3 days.
                    """
                    after_3_days = datetime.datetime.today() + datetime.timedelta(days=3)

                    warning_overage_emails.apply_async(args=[subscriber, plan_name, total_submissions, extra_submissions_charge,
                                                             renewal_date, email], eta=after_3_days)

                elif length_of_day.days > 7:
                    """
                        Send email after 7 days if upcoming renewal date is more than 7 days.
                    """
                    after_one_week = datetime.datetime.today() + datetime.timedelta(days=7)
                    warning_overage_emails.apply_async(args=[subscriber, plan_name, total_submissions, extra_submissions_charge,
                                                             renewal_date, email], eta=after_one_week)

                elif period_type == "Year" and length_of_day.days > 30:
                    """
                        Send email after 30 days if upcoming renewal date is more than 30 days.
                    """
                    after_one_month = datetime.datetime.today() + datetime.timedelta(days=30)
                    warning_overage_emails_monthly.apply_async(
                        args=[subscriber, plan_name, total_submissions, extra_submissions_charge, renewal_date, email],
                        eta=after_one_month)
import calendar
from django.utils import timezone

def sync_form(fxf):
    #sync
    create_async_export(xform, export_type, query, force_xlsx, options, is_project, id, site_id, version, sync_to_gsuit, user)

    pass

def sync_form_controller(sync_type, sync_day, fxf, month_days):
    date = timezone.date()
    if sync_type == Project.MONTHLY:
        if date.day == sync_day or sync_day > month_days:
            sync_form(fxf)

    elif sync_type == project.FORTNIGHT:
        pass

    elif sync_type == WEEKLY:
        if (((date.weekday() + 1) % 7) + 1) == sync_day:
            sync_form(fxf)
    else:
        sync_form(fxf)


@shared_task(time_limit=300, soft_time_limit=300)
def update_sites_progress(pk, task_id):
    try:
        time.sleep(3)
        obj = ProgressSettings.objects.get(pk=pk)
        project = obj.project
        total_sites = project.sites.filter(is_active=True,
                                           enable_subsites=False).count()
        page_size = 1000
        page = 0
        while total_sites > 0:
            sites = project.sites.filter(is_active=True, enable_subsites=False)[
                    page*page_size:(page+1)*page_size]
            print("updating site progress batch in project ", project.id, page*page_size, (page+1)*page_size)
            for site in sites:
                set_site_progress(site, project, obj)
            total_sites -= page_size
            page += 1
        CeleryTaskProgress.objects.filter(id=task_id).update(status=2)
    except ProgressSettings.DoesNotExist:
        CeleryTaskProgress.objects.filter(id=task_id).update(status=3)


def scheduled_gsuit_sync():
    month_days = calendar.monthrange(now.year, now.month)[1]
    projects = Project.objects.filter(is_active=True).exclude(gsuit_sync=Project.MANUAL)

    for project in projects:
        #generate reports
        for fxf in project.project_forms.exclude(sync_schedule__schedule=Project.MANUAL):
            if fxf.sync_schedule:
                sync_form(fxf.sync_schedule.schedule, fxf.sync_schedule.day, fxf, month_days)
            else:
                sync_form_controller(project.gsuit_sync, project.gsuit_sync_day, fxf, month_days)
        project.gsuit_sync


@shared_task(time_limit=900, max_retries=2, soft_time_limit=900)
def create_site_meta_attribs_ans_history(pk, task_id):
    from onadata.apps.fieldsight.utils.siteMetaAttribs import get_site_meta_ans
    total_sites = Site.objects.filter(is_active=True, project=pk).count()
    page_size = 1000
    page = 0
    try:
        while total_sites > 0:
            sites = Site.objects.filter(is_active=True, project=pk)[
                    page * page_size:(page + 1) * page_size]
            print("updating site Metas batch for project ", pk, page * page_size, (page + 1) * page_size)
            for site in sites:
                metas = get_site_meta_ans(site.id)
                if metas == site.site_meta_attributes_ans:
                    continue
                else:
                    SiteMetaAttrAnsHistory.objects.create(site=site, meta_attributes_ans=site.site_meta_attributes_ans,
                                                          status=2)
                    site.site_meta_attributes_ans = metas
                    site.save()
            total_sites -= page_size
            page += 1
            CeleryTaskProgress.objects.filter(id=task_id).update(status=2)
    except Exception:
        CeleryTaskProgress.objects.filter(id=task_id).update(status=3)


def get_submission_answer_by_question(sub_answers={}, question_name=""):
    answer = sub_answers.get(question_name, None)
    if not answer:
        for k, v in sub_answers.items():
            if isinstance(v, list):
                return get_submission_answer_by_question(v[0], question_name)
    return answer


@shared_task(max_retries=5, time_limit=300, soft_time_limit=300)
def update_meta_details(fs_proj_xf_id, instance_id, task_id, site_id):
    try:
        instance = Instance.objects.get(id=instance_id)
        fs_proj_xf = FieldSightXF.objects.get(id=fs_proj_xf_id)
        site = Site.objects.get(id=int(site_id))
        site_picture = site.project.site_basic_info.get('site_picture', None)
        if site_picture and site_picture.get('question_type', '') == 'Form' and \
                site_picture.get('form_id', 0) == fs_proj_xf.id and site_picture.get(
                'question', {}):
            print("has site logo settings and matches form ")
            question_name = site_picture['question'].get('name', '')
            logo_url = get_submission_answer_by_question(instance.json, question_name)
            if logo_url:
                attachment = Attachment.objects.get(instance=instance, media_file_basename=logo_url)
                site.logo = attachment.media_file

        site_loc = fs_proj_xf.project.site_basic_info.get('site_location', None)
        if site_loc and site_loc.get('question_type', '') == 'Form' and site_loc.get('form_id', 0) == fs_proj_xf.id and site_loc.get('question', {}):
            question_name = site_loc['question'].get('name', '')
            location = get_submission_answer_by_question(instance.json, question_name)
            if location:
                location_float = list(map(lambda x: float(x), str(location).split(' ')))
                site.location = Point(round(float(location_float[1]), 6), round(float(location_float[0]), 6), srid=4326)

        for featured_img in fs_proj_xf.project.site_featured_images:
            if featured_img.get('question_type', '') == 'Form' and featured_img.get('form_id', '') == str(fs_proj_xf.id) and featured_img.get('question', {}):

                question_name = featured_img['question'].get('name', '')
                logo_url = instance.json.get(question_name)
                if logo_url:
                    attachments = {}
                    attachment = Attachment.objects.get(instance=instance, media_file_basename=logo_url)
                    attachments['_attachments'] = attachment.media_file.url
                    attachments['_id'] = instance.id
                    site.site_featured_images[question_name] = attachments
        site.save()

        # change site meta attributes answer
        meta_ans = site.site_meta_attributes_ans
        for item in fs_proj_xf.project.site_meta_attributes:
            if item.get('question_type') == 'Form' and fs_proj_xf.id == item.get('form_id', 0):
                if item['question']['type'] == "repeat":
                    answer = ""
                else:
                    answer = instance.json.get(item.get('question').get('name'), '')
                if item['question']['type'] in ['photo', 'video', 'audio'] and answer is not "":
                    answer = 'http://app.fieldsight.org/attachment/medium?media_file=' + fs_proj_xf.xf.user.username + '/attachments/' + answer
                meta_ans[item['question_name']] = answer

            elif item.get('question_type') == 'FormSubStat' and fs_proj_xf.id == item.get('form_id', 0):
                if instance.date_modified:
                    answer = "Last submitted on " + instance.date_modified.strftime("%d %b %Y %I:%M %P")
                else:
                    answer = "Last submitted on " + instance.date_created.strftime("%d %b %Y %I:%M %P")

                meta_ans[item['question_name']] = answer

            elif item.get('question_type') == "FormQuestionAnswerStatus":
                get_answer = instance.json.get(item.get('question').get('name'), None)
                if get_answer:
                    answer = "Answered"
                else:
                    answer = ""
                meta_ans[item['question_name']] = answer

            elif item.get('question_type') == "FormSubCountQuestion":
                meta_ans[item['question_name']] = fs_proj_xf.project_form_instances.filter(site_id=site.id).count()
        if meta_ans != site.site_meta_attributes_ans:
            SiteMetaAttrAnsHistory.objects.create(site=site, meta_attributes_ans=site.site_meta_attributes_ans, status=1)
            site.site_meta_attributes_ans = meta_ans
            site.save()
        CeleryTaskProgress.objects.filter(id=task_id).update(status=2)
    except Exception as e:
        print('Exception occured', e)
        CeleryTaskProgress.objects.filter(id=task_id).update(status=3)

@shared_task()
def update_current_progress_site(site_id):
    site = Site.objects.get(pk=site_id)
    set_site_progress(site, site.project)
    try:
        status = site.site_instances.order_by('-date').first().form_status
    except:
        status = 0
    site.current_status = status
    site.save()


def update_basic_info_in_site(pk, location_changed, picture_changed, location_form, location_question, picture_form,
                              picture_question, site):
    if location_changed:
        submission = FInstance.objects.filter(
            site=site, project_fxf__id=location_form).order_by('-date').first()
        if submission:
            location = get_submission_answer_by_question(submission.instance.json, location_question)
            if location:
                location_float = list(map(lambda x: float(x), str(location).split(' ')))
                site.location = Point(round(float(location_float[1]), 6), round(float(location_float[0]), 6), srid=4326)

    if picture_changed:
        submission = FInstance.objects.filter(
            site=site, project_fxf__id=picture_form).order_by('-date').first()
        if submission:
            logo_url = get_submission_answer_by_question(submission.instance.json, picture_question)
            if logo_url:
                attachment = Attachment.objects.get(instance=submission.instance, media_file_basename=logo_url)
                site.logo = attachment.media_file
    site.save()


@shared_task(time_limit=900, max_retries=2, soft_time_limit=900)
def update_sites_info(pk, location_changed, picture_changed,
                      location_form, location_question, picture_form, picture_question):
    total_sites = Site.objects.filter(is_active=True, project=pk).count()
    page_size = 1000
    page = 0
    try:
        while total_sites > 0:
            sites = Site.objects.filter(is_active=True, project=pk)[
                    page * page_size:(page + 1) * page_size]
            print("updating site Metas batch for project ", pk, page * page_size, (page + 1) * page_size)
            for site in sites:
                update_basic_info_in_site(pk, location_changed,
                                          picture_changed, location_form,
                                          location_question, picture_form, picture_question, site)

            total_sites -= page_size
            page += 1
    except Exception as e:
        print(str(e))
